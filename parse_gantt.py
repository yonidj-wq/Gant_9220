#!/usr/bin/env python3
"""
Parse גאנט.xlsx and write gantt-site/data/events.json.
Run from the folder that contains both גאנט.xlsx and gantt-site/.
"""
import sys
import io
import os
import datetime
import json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("Please install openpyxl:  pip install openpyxl")
    sys.exit(1)

XLSX_FILE   = 'גאנט.xlsx'
# Support both local layout (gantt-site/data/) and GitHub repo layout (data/)
_script_dir = os.path.dirname(os.path.abspath(__file__))
if os.path.isdir(os.path.join(_script_dir, 'gantt-site', 'data')):
    OUTPUT_FILE = os.path.join(_script_dir, 'gantt-site', 'data', 'events.json')
else:
    OUTPUT_FILE = os.path.join(_script_dir, 'data', 'events.json')

# Theme index → category name (based on Google Sheets / Office accent colours)
THEME_CAT = {
    7: 'גדודי',    # accent4 = green
    6: 'פלוגתי',   # accent3 = yellow/amber
    8: 'חטיבתי',   # accent5 = orange
    9: 'נוספים',   # accent6 = teal
    5: 'קיפול',    # accent2 = red
    4: 'כללי',     # accent1 = blue (headers)
}
RGB_CAT = {
    'FFF1C232': 'פלוגתי',   # amber/yellow fill used for company training
}

CATEGORIES = [
    {'name': 'גדודי',  'color': '#34A853'},
    {'name': 'פלוגתי', 'color': '#F9AB00'},
    {'name': 'חטיבתי', 'color': '#FF6D01'},
    {'name': 'נוספים', 'color': '#46BDC6'},
    {'name': 'קיפול',  'color': '#EA4335'},
    {'name': 'ג״מ',    'color': '#7B68EE'},
    {'name': 'כללי',   'color': '#4285F4'},
]


def get_category(cell):
    fill = cell.fill
    if not fill or fill.fill_type != 'solid':
        return 'כללי'
    fg = fill.fgColor
    if fg.type == 'theme':
        return THEME_CAT.get(fg.theme, 'כללי')
    elif fg.type == 'rgb':
        rgb = fg.rgb
        if rgb and rgb not in ('00000000', 'FFFFFFFF', 'FF000000'):
            return RGB_CAT.get(rgb, 'כללי')
    return 'כללי'


def parse(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['גאנט']

    # Find date rows: rows where column B contains a datetime
    date_rows = {}
    for row_idx in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row_idx, column=2)
        if isinstance(cell_b.value, datetime.datetime):
            dates = {}
            for c in range(2, 9):   # columns B–H
                cell = ws.cell(row=row_idx, column=c)
                if isinstance(cell.value, datetime.datetime):
                    dates[c] = cell.value.date()
            if dates:
                date_rows[row_idx] = dates

    sorted_date_rows = sorted(date_rows.keys())

    def find_week_for_row(event_row):
        prev = None
        for dr in sorted_date_rows:
            if dr <= event_row:
                prev = dr
            else:
                break
        return prev

    events = []

    for row_idx in range(2, ws.max_row + 1):  # skip row 1 (day-letter headers)
        for col in range(2, 10):               # columns B–I
            cell = ws.cell(row=row_idx, column=col)
            val  = cell.value
            if val is None or isinstance(val, datetime.datetime):
                continue

            cat      = get_category(cell)
            week_row = find_week_for_row(row_idx)
            if not week_row:
                continue

            if col == 9:   # column I = ג"מ — spans the whole week
                dates  = date_rows[week_row]
                cols   = sorted(dates.keys())
                start  = dates[cols[0]]
                end    = dates[cols[-1]]
                events.append({
                    'id':       len(events) + 1,
                    'name':     str(val),
                    'category': 'ג״מ',
                    'start':    start.isoformat(),
                    'end':      end.isoformat(),
                    'notes':    '',
                })
            else:
                if col in date_rows[week_row]:
                    date = date_rows[week_row][col]
                    events.append({
                        'id':       len(events) + 1,
                        'name':     str(val),
                        'category': cat,
                        'start':    date.isoformat(),
                        'end':      date.isoformat(),
                        'notes':    '',
                    })

    events.sort(key=lambda e: (e['start'], e['category']))
    return events


def main():
    if not os.path.exists(XLSX_FILE):
        print(f"Error: {XLSX_FILE} not found in current directory.")
        sys.exit(1)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    events = parse(XLSX_FILE)

    output = {'events': events, 'categories': CATEGORIES}
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Parsed {len(events)} events → {OUTPUT_FILE}")
    for e in events:
        print(f"  {e['start']} | {e['category']:8s} | {e['name']}")


if __name__ == '__main__':
    main()
